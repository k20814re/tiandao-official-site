#!/usr/bin/env python3
"""Convert the official TianDao content workbook into browser-ready JSON."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "TianDao_Website_Content.xlsx"
OUTPUT_PATH = ROOT / "data" / "site-content.json"

EXPECTED_HEADERS = {
    "首頁設定": ["Key", "值", "用途"],
    "世界觀設定": ["Key", "值", "用途"],
    "網站文字": ["Key", "值", "用途"],
    "網站素材": ["Key", "值", "用途"],
    "最新消息": ["啟用", "日期", "標題", "說明", "圖片路徑", "標章"],
    "開發進度": ["啟用", "階段", "標題", "說明", "目前階段"],
    "社群連結": ["啟用", "平台", "狀態文字", "網址"],
    "世界歷程": ["啟用", "排序", "名稱", "目前階段"],
    "修煉境界": ["啟用", "排序", "名稱"],
    "宗門流派": ["啟用", "排序", "名稱"],
    "世界地圖": ["啟用", "排序", "名稱"],
    "遊戲特色": ["啟用", "排序", "標題", "說明"],
    "App標籤": ["啟用", "排序", "名稱"],
    "畫廊內容": ["啟用", "排序", "版型", "圖片路徑", "替代文字", "圖說"],
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def enabled(value: Any) -> bool:
    return clean(value).lower() in {"是", "true", "1", "yes", "y"}


def number(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def date_text(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = clean(value)
    if not text:
        return ""
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"無法辨識日期：{text}")


def rows_for(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"找不到工作表：{sheet_name}")

    sheet = workbook[sheet_name]
    expected = EXPECTED_HEADERS[sheet_name]
    actual = [clean(sheet.cell(row=3, column=index).value) for index in range(1, len(expected) + 1)]
    if actual != expected:
        raise ValueError(
            f"{sheet_name} 第 3 列欄位不可更名。預期 {expected}，目前為 {actual}"
        )

    result: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=4, max_col=len(expected), values_only=True):
        if not any(value is not None and clean(value) for value in values):
            continue
        result.append(dict(zip(expected, values)))
    return result


def optional_rows_for(workbook, sheet_name: str) -> list[dict[str, Any]]:
    """Read a newly-added sheet without breaking older workbooks."""
    if sheet_name not in workbook.sheetnames:
        return []
    return rows_for(workbook, sheet_name)


def key_value_rows(workbook, sheet_name: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in optional_rows_for(workbook, sheet_name):
        key = clean(row["Key"])
        if key:
            result[key] = clean(row["值"])
    return result


def build_content(workbook) -> dict[str, Any]:
    settings: dict[str, str] = {}
    for row in rows_for(workbook, "首頁設定"):
        key = clean(row["Key"])
        if key:
            settings[key] = clean(row["值"])

    world_settings: dict[str, str] = {}
    for row in rows_for(workbook, "世界觀設定"):
        key = clean(row["Key"])
        if key:
            world_settings[key] = clean(row["值"])

    site_text = key_value_rows(workbook, "網站文字")
    site_media = key_value_rows(workbook, "網站素材")

    news = [
        {
            "date": date_text(row["日期"]),
            "title": clean(row["標題"]),
            "description": clean(row["說明"]),
            "image": clean(row["圖片路徑"]),
            "badge": clean(row["標章"]),
        }
        for row in rows_for(workbook, "最新消息")
        if enabled(row["啟用"]) and clean(row["標題"])
    ]
    news.sort(key=lambda item: item["date"], reverse=True)

    roadmap = [
        {
            "phase": clean(row["階段"]),
            "title": clean(row["標題"]),
            "description": clean(row["說明"]),
            "current": enabled(row["目前階段"]),
        }
        for row in rows_for(workbook, "開發進度")
        if enabled(row["啟用"]) and clean(row["標題"])
    ]

    social = [
        {
            "platform": clean(row["平台"]),
            "status": clean(row["狀態文字"]),
            "url": clean(row["網址"]),
        }
        for row in rows_for(workbook, "社群連結")
        if enabled(row["啟用"]) and clean(row["平台"])
    ]

    world_flow = [
        {
            "order": number(row["排序"]),
            "name": clean(row["名稱"]),
            "current": enabled(row["目前階段"]),
        }
        for row in rows_for(workbook, "世界歷程")
        if enabled(row["啟用"]) and clean(row["名稱"])
    ]
    world_flow.sort(key=lambda item: item["order"])

    realms = [
        {"order": number(row["排序"]), "name": clean(row["名稱"])}
        for row in rows_for(workbook, "修煉境界")
        if enabled(row["啟用"]) and clean(row["名稱"])
    ]
    realms.sort(key=lambda item: item["order"])

    sect_tags = [
        {"order": number(row["排序"]), "name": clean(row["名稱"])}
        for row in rows_for(workbook, "宗門流派")
        if enabled(row["啟用"]) and clean(row["名稱"])
    ]
    sect_tags.sort(key=lambda item: item["order"])

    map_items = [
        {"order": number(row["排序"]), "name": clean(row["名稱"])}
        for row in rows_for(workbook, "世界地圖")
        if enabled(row["啟用"]) and clean(row["名稱"])
    ]
    map_items.sort(key=lambda item: item["order"])

    features = [
        {
            "order": number(row["排序"]),
            "title": clean(row["標題"]),
            "description": clean(row["說明"]),
        }
        for row in optional_rows_for(workbook, "遊戲特色")
        if enabled(row["啟用"]) and clean(row["標題"])
    ]
    features.sort(key=lambda item: item["order"])

    app_tags = [
        {"order": number(row["排序"]), "name": clean(row["名稱"])}
        for row in optional_rows_for(workbook, "App標籤")
        if enabled(row["啟用"]) and clean(row["名稱"])
    ]
    app_tags.sort(key=lambda item: item["order"])

    gallery = [
        {
            "order": number(row["排序"]),
            "layout": clean(row["版型"]).lower() or "normal",
            "image": clean(row["圖片路徑"]),
            "alt": clean(row["替代文字"]),
            "caption": clean(row["圖說"]),
        }
        for row in optional_rows_for(workbook, "畫廊內容")
        if enabled(row["啟用"]) and clean(row["圖片路徑"])
    ]
    gallery.sort(key=lambda item: item["order"])

    world = {
        "kicker": world_settings.get("worldKicker", ""),
        "title": world_settings.get("worldTitle", ""),
        "subtitle": world_settings.get("worldSubtitle", ""),
        "realm": {
            "label": world_settings.get("realmLabel", ""),
            "title": world_settings.get("realmTitle", ""),
            "description": world_settings.get("realmDescription", ""),
        },
        "sect": {
            "label": world_settings.get("sectLabel", ""),
            "title": world_settings.get("sectTitle", ""),
            "description": world_settings.get("sectDescription", ""),
        },
        "map": {
            "label": world_settings.get("mapLabel", ""),
            "title": world_settings.get("mapTitle", ""),
            "description": world_settings.get("mapDescription", ""),
        },
    }

    return {
        "schemaVersion": 3,
        "settings": settings,
        "text": site_text,
        "media": site_media,
        "news": news,
        "roadmap": roadmap,
        "social": social,
        "features": features,
        "appTags": app_tags,
        "gallery": gallery,
        "world": world,
        "worldFlow": world_flow,
        "realms": realms,
        "sectTags": sect_tags,
        "mapItems": map_items,
    }


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"找不到 {WORKBOOK_PATH.name}")

    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    try:
        content = build_content(workbook)
    finally:
        workbook.close()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(content, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"已更新 {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
